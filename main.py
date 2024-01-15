import os
import openpyxl
from xhtml2pdf import pisa
from time import sleep

def exitePlanilha(nomeArquivo):
    if not os.path.isfile(nomeArquivo):
        workbook = openpyxl.Workbook()
        planilha = workbook.active
        planilha.title = 'Planilha1'
        planilha.append(['Nome', 'Status', 'Sistema Operacional', 'Chave de Ativação', 'Partições'])
        workbook.save(nomeArquivo)
        print(f'Arquivo {nomeArquivo} criado com sucesso!')
        return planilha
    else:
        workbook = openpyxl.load_workbook(nomeArquivo)
        planilha = ['Planilha1']
        print("Base de dados importada com sucesso!")
        return planilha
        
def menu():
    print("="*32 + " Gerenciador de dispositivos locais - Prominas " + "="*32)
    print("1. Nova máquina.\n2. Mostrar todos.\n3. Mostrar uma máquina.\n4. Alterar especificações de uma máquina.\n5. Excluir.\n6. Excluir todos.\n7. Gerar PDF.\n8. Sair")
    print("="*80)
    opcao = input("Entre com o dígito da função desejada: ")
    return opcao

def validaIdentificador(planilhaDeDispositivos):
    while True:
        y = input("Nome da máquina: ").upper()
        for x in planilhaDeDispositivos:
            if x[0].value == y:
                print("Essa máquina já existe. Tente novamente.")
                break
        else:
            return y

def incluirMaquina(planilhaDeDispositivos):
    print("="*32 + " Incluindo Dispositivo " + "="*32)
    nova_linha = [None] * 5
    valida = validaIdentificador(planilhaDeDispositivos)
    nova_linha[0] = valida
    exiSta = False
    while not exiSta:
        x = input("Status da máquina (Ativo/Inativo): ").lower().capitalize()
        if x in ["Ativo", "Inativo"]:
            exiSta = True
            nova_linha[1] = x
    nova_linha[2] = input("Sistema Operacional: ").lower().capitalize()
    nova_linha[3] = input("Chave de ativação: ").upper()
    
    particoes = []
    y = True
    while y:
        particao = input("Espaço de armazenamento da partição: ").upper()
        particoes.append(particao)
        pgt = input("Deseja inserir uma nova partição (s/n): ").lower()
        if pgt == "n":
            y = False
    
    nova_linha[4] = ', '.join(particoes)
    planilhaDeDispositivos.append(nova_linha)
    print("Dispositivo inserido com sucesso!")
    planilhaDeDispositivos.parent.save('Dispositivos prmns.xlsx')

def mostrarTodos(planilhaDeDispositivos):
    print("="*32 + " Exibindo dispositivos " + "="*32)
    if planilhaDeDispositivos.max_row > 1:
        for linha in planilhaDeDispositivos.iter_rows(min_row = 2):
            print(f"{linha[0].value} | {linha[1].value} | {linha[2].value} | {linha[3].value} | [{linha[4].value}]")
    else:
        print("Não existem dispositivos cadastrados na base de dados atual. Faça a inclusão de novos dispositivos para utilizar a função mostra todos.")

def buscarDispositivo(planilhaDeDispositivos):
    x = input("Nome da máquina: ").upper()
    for linha in planilhaDeDispositivos.iter_rows(min_row = 2):
        if linha[0].value == x:
            return linha
    return False

def imprimirMaquinaEspecifica(planilhaDeDispositivos):
    print("="*32 + "Mostrando dispositivo específico" + "="*32)
    indDisp = buscarDispositivo(planilhaDeDispositivos)
    if indDisp:
        print(f"{indDisp[0].value} | {indDisp[1].value} | {indDisp[2].value} | {indDisp[3].value} | [{indDisp[4].value}]")
    else:
        print("O dispositivo indicado não podê ser encontrado. Verifique se o mesmo consta na base de dados atual.")

def alterarEspecificacao(planilhaDeDispositivos):
    indDisp = buscarDispositivo(planilhaDeDispositivos)
    if indDisp:
        opcao = input("="*32 + " Alterando especificações do dispositivo " + "="*32 + "\n1. Status.\n2. Sistema Operacional.\n3. Chave de ativação.\n4. Partições\nEntre com o dígito da opção que deseja alterar no dispostivo: ")
        if opcao == "1":
            print("Alterando Status do dispositivo\n")
            exiSta = False
            while not exiSta:
                x = input("Status da máquina (Ativo/Inativo): ").lower().capitalize()
                if x in ["Ativo", "Inativo"]:
                    exiSta = True
                    indDisp[1].value = x
        elif opcao == "2":
            print("Alterando Sistema Operacional do dispositivo\n")
            indDisp[2].value = input("Sistema Operacional: ").lower().capitalize()
        elif opcao == "3":
            print("Alterando chave de ativação do dispositivo\n")
            indDisp[3].value = input("Chave de ativação: ").upper()
        elif opcao == "4":
            print("Alterando partições do dispositivo: ")
            i = True
            x = []
            while i:
                particao = input("Espaço de armazenamento da partição: ").upper()
                x.append(f"{particao}")
                pgt = input("Deseja inserir uma nova partição (s/n): ").lower()
                if pgt == "n":
                    i = False
            indDisp[4].value = x
        else:
            print("Opção inválida. Tente novamente.")
    else:
        print("O dispositivo indicado não podê ser encontrado. Verifique se o mesmo consta na base de dados atual.")

def excluirDispositivo(planilhaDeDispositivos):
    print("="*32 + " Excluindo dispositivo " + "="*32)
    indDisp = buscarDispositivo(planilhaDeDispositivos)
    if (indDisp) and (indDisp[1].value != 'Ativo'):
        planilhaDeDispositivos.delete_rows(indDisp[0].row)
        print("Dispositivo removido com sucesso!")
    else:
        print("O dispositivo indicado não podê ser encontrado ou não pode ser excluído devido seu status. Verifique se o mesmo consta na base de dados atual e qual e se seu status está inativado.")

def excluirTodos(planilhaDeDispositivos):
    print("="*32 + " Excluindo Todos os dispositivos Inativos " + "="*32)
    for linha in planilhaDeDispositivos.iter_rows(min_row = 2):
        if linha[1].value != "Ativo":
            planilhaDeDispositivos.delete_rows(linha[0].row)
    print("Dispositivos removidos com sucesso!")

def gerarPDF(planilhaDeDispositivos):
    print("="*32 + "Gerando PDF" + "="*32)
    pdf_file = "Etiquetas prmns.pdf"

    # Criar um documento HTML
    html_content = f"""
    <!DOCTYPE html>
    <html lang="pt-br">
    <head>
        <style>
            body {{
                font-family: 'Arial Narrow', Arial, sans-serif;
            }}
            .container {{
                display: flex;
                flex-wrap: wrap;
            }}
            table {{
                width: 100%;
            }}
            th, td {{
                border: 1px solid black;
                text-align: center;
                padding: 8px;
                font-size: medium;
            }}
            th {{
                background-color: #f2f2f2;
                text-align: center;
            }}
            tr > td{{
                font-size: 9pt;
            }}
        </style>
    </head>
    <body>
        <div class="container">
    """

    # Adicionar informações de cada máquina ao HTML
    for linha in planilhaDeDispositivos.iter_rows(min_row = 2):    
        html_content += f"""
        <div class="column">
            <table>
                <tr>
                    <th><h1>Nome do dispositivo</h1></th>
                    <th><h1>Sistema Operacional</h1></th>
                    <th><h1>Chave de Ativação</h1></th>
                </tr>
                <tr>
                    <td>{linha[0].value}</td>
                    <td>{linha[2].value}</td>
                    <td>{linha[3].value}</td>
                </tr>
            </table>
        </div>
        """
                    
    # Fechar o HTML
    html_content += """
        </div>
    </body>
    </html>
    """

    # Criar o PDF a partir do HTML
    with open(pdf_file, 'w+b') as pdf:
        pisa.CreatePDF(html_content, dest=pdf)

    print(f"PDF gerado com sucesso em {pdf_file}")

def desligandoSistema():
    for x in range(1,4):
        print(f"Desligando{'.'*x}", end='\r')
        sleep(0.5)

def carregarArquivo():
    print("="*32 + " Gerenciador de dispositivos locais - Prominas " + "="*32)
    nomeArquivo = input("Entre com o nome da base de dados: ").lower().capitalize()
    nomeArquivo += '.xlsx'
    return nomeArquivo

def main():
    arquivo = carregarArquivo()
    verifArq = exitePlanilha(arquivo)
    opcao = menu()
    while opcao != "8":
        if opcao == "1":
            incluirMaquina(verifArq)
        elif opcao == "2":
            mostrarTodos(verifArq)
        elif opcao == "3":
            imprimirMaquinaEspecifica(verifArq)
        elif opcao == "4":
            alterarEspecificacao(verifArq)
        elif opcao == "5":
            excluirDispositivo(verifArq)
        elif opcao == "6":
            excluirTodos(verifArq)
        elif opcao == "7":
            gerarPDF(verifArq)
        else:
            print("Opção inválida. Tente novamente com uma das opções disponíveis no menu.")
        opcao = menu()
        verifArq.parent.save('Dispositivos prmns.xlsx')
    desligandoSistema()

if __name__ == "__main__":
    main()